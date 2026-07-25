# routes/supplier.py - COMPLETE WITH PROFESSIONAL INVOICE-STYLE PDF, SHOP NAME DISPLAY, FULL BIDIRECTIONAL STATUS CHANGES & EXCEL EXPORT
from flask import request, jsonify, current_app
from flask_login import login_required, current_user
from extensions import db
from models.supplier import Supplier
from models.inventory import Product
from models.sales import Sale
from models.purchase_order import PurchaseOrder, PurchaseOrderItem, OrderStatusHistory
from models.shop import Shop
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

def get_current_shop_id():
    """Get the current shop ID from the logged-in user"""
    if hasattr(current_user, 'id'):
        return current_user.id
    return None

def get_shop_name(shop_id):
    """Get shop name by ID, returns formatted string if not found"""
    if not shop_id:
        return 'System'
    shop = Shop.query.filter_by(id=shop_id).first()
    if shop:
        return shop.name
    return f'Shop #{shop_id}'

def get_available_status_transitions(current_status):
    """
    Get available status transitions for an order.
    Supports reverting to previous statuses to fix mistakes.
    """
    transitions = {
        'Pending': ['Ordered', 'Cancelled'],
        'Ordered': ['Received', 'Cancelled', 'Pending'],
        'Received': ['Pending', 'Ordered'],
        'Cancelled': ['Pending', 'Ordered']
    }
    return transitions.get(current_status, [])

def validate_status_transition(current_status, new_status):
    """Validate if a status transition is allowed"""
    allowed = get_available_status_transitions(current_status)
    return new_status in allowed

def enrich_order_with_shop_name(order_dict, order):
    """Add created_by_name to order dict"""
    if order.created_by:
        shop = Shop.query.filter_by(id=order.created_by).first()
        if shop:
            order_dict['created_by_name'] = shop.name
        else:
            order_dict['created_by_name'] = f'Shop #{order.created_by}'
    else:
        order_dict['created_by_name'] = 'System'
    return order_dict

def enrich_history_with_shop_name(history_dict, history):
    """Add changed_by_name to history dict"""
    if history.changed_by:
        shop = Shop.query.filter_by(id=history.changed_by).first()
        if shop:
            history_dict['changed_by_name'] = shop.name
        else:
            history_dict['changed_by_name'] = f'Shop #{history.changed_by}'
    else:
        history_dict['changed_by_name'] = 'System'
    return history_dict

def init_supplier_routes(app):
    
    # ============ SUPPLIER ROUTES ============
    
    @app.route('/api/suppliers', methods=['GET'])
    @login_required
    def get_suppliers():
        """Get all suppliers for the current shop only"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            status = request.args.get('status')
            item_name = request.args.get('item_name')
            search = request.args.get('search')
            
            query = Supplier.query.filter_by(shop_id=shop_id)
            
            if status:
                query = query.filter_by(status=status)
            if item_name:
                query = query.filter_by(item_name=item_name)
            if search:
                query = query.filter(
                    db.or_(
                        Supplier.name.contains(search),
                        Supplier.contact_person.contains(search),
                        Supplier.email.contains(search),
                        Supplier.phone.contains(search)
                    )
                )
            
            suppliers = query.order_by(Supplier.created_at.desc()).all()
            return jsonify([s.to_dict() for s in suppliers])
            
        except Exception as e:
            logger.error(f"Error fetching suppliers: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers/<int:supplier_id>', methods=['GET'])
    @login_required
    def get_supplier(supplier_id):
        """Get a specific supplier for the current shop"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            supplier = Supplier.query.filter_by(id=supplier_id, shop_id=shop_id).first()
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            
            return jsonify(supplier.to_dict())
            
        except Exception as e:
            logger.error(f"Error fetching supplier: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers', methods=['POST'])
    @login_required
    def create_supplier():
        """Create a new supplier for the current shop"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            data = request.get_json()
            
            required = ['name', 'contact_person', 'email', 'phone']
            missing = [f for f in required if f not in data]
            if missing:
                return jsonify({'error': f'Missing required fields: {", ".join(missing)}'}), 400
            
            if Supplier.query.filter_by(email=data['email'], shop_id=shop_id).first():
                return jsonify({'error': 'Supplier with this email already exists in your shop'}), 400
            
            supplier = Supplier(
                shop_id=shop_id,
                name=data['name'],
                contact_person=data['contact_person'],
                email=data['email'],
                phone=data['phone'],
                address=data.get('address', ''),
                city=data.get('city', ''),
                state=data.get('state', ''),
                country=data.get('country', ''),
                postal_code=data.get('postal_code', ''),
                item_name=data.get('item_name', ''),
                status=data.get('status', 'Active'),
                rating=data.get('rating', 0.0),
                notes=data.get('notes', '')
            )
            
            db.session.add(supplier)
            db.session.commit()
            
            return jsonify(supplier.to_dict()), 201
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error creating supplier: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers/<int:supplier_id>', methods=['PUT'])
    @login_required
    def update_supplier(supplier_id):
        """Update a supplier for the current shop"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            supplier = Supplier.query.filter_by(id=supplier_id, shop_id=shop_id).first()
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            
            data = request.get_json()
            
            if 'name' in data:
                supplier.name = data['name']
            if 'contact_person' in data:
                supplier.contact_person = data['contact_person']
            if 'email' in data:
                existing = Supplier.query.filter_by(email=data['email'], shop_id=shop_id).first()
                if existing and existing.id != supplier_id:
                    return jsonify({'error': 'Email already in use'}), 400
                supplier.email = data['email']
            if 'phone' in data:
                supplier.phone = data['phone']
            if 'address' in data:
                supplier.address = data['address']
            if 'city' in data:
                supplier.city = data['city']
            if 'state' in data:
                supplier.state = data['state']
            if 'country' in data:
                supplier.country = data['country']
            if 'postal_code' in data:
                supplier.postal_code = data['postal_code']
            if 'item_name' in data:
                supplier.item_name = data['item_name']
            if 'status' in data:
                supplier.status = data['status']
            if 'rating' in data:
                supplier.rating = float(data['rating'])
            if 'notes' in data:
                supplier.notes = data['notes']
            
            db.session.commit()
            return jsonify(supplier.to_dict())
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error updating supplier: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers/<int:supplier_id>', methods=['DELETE'])
    @login_required
    def delete_supplier(supplier_id):
        """Delete a supplier for the current shop (soft delete - mark as inactive)"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            supplier = Supplier.query.filter_by(id=supplier_id, shop_id=shop_id).first()
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            
            supplier.status = 'Inactive'
            db.session.commit()
            
            return jsonify({'message': 'Supplier deactivated successfully'})
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error deleting supplier: {e}")
            return jsonify({'error': str(e)}), 500
    
    # ============ SUPPLIER STATS ============
    
    @app.route('/api/suppliers/stats', methods=['GET'])
    @login_required
    def get_supplier_stats():
        """Get supplier statistics for the current shop only"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            suppliers = Supplier.query.filter_by(shop_id=shop_id).all()
            active = Supplier.query.filter_by(shop_id=shop_id, status='Active').count()
            inactive = Supplier.query.filter_by(shop_id=shop_id, status='Inactive').count()
            pending = Supplier.query.filter_by(shop_id=shop_id, status='Pending').count()
            
            total_products = sum(s.total_products for s in suppliers)
            total_orders = sum(s.total_orders for s in suppliers)
            total_spent = sum(s.total_spent for s in suppliers)
            
            category_breakdown = {}
            for supplier in suppliers:
                if supplier.item_name:
                    if supplier.item_name not in category_breakdown:
                        category_breakdown[supplier.item_name] = 0
                    category_breakdown[supplier.item_name] += 1
            
            return jsonify({
                'total': len(suppliers),
                'active': active,
                'inactive': inactive,
                'pending': pending,
                'total_products': total_products,
                'total_orders': total_orders,
                'total_spent': float(total_spent),
                'category_breakdown': category_breakdown
            })
            
        except Exception as e:
            logger.error(f"Error fetching supplier stats: {e}")
            return jsonify({'error': str(e)}), 500
    
    # ============ UPDATE SUPPLIER STATS ============
    
    @app.route('/api/suppliers/update-stats', methods=['POST'])
    @login_required
    def update_supplier_stats():
        """Update all supplier stats from products for the current shop only"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            suppliers = Supplier.query.filter_by(shop_id=shop_id).all()
            updated_count = 0
            
            for supplier in suppliers:
                products = Product.query.filter_by(supplier_id=supplier.id, shop_id=shop_id).all()
                supplier.total_products = len(products)
                updated_count += 1
            
            db.session.commit()
            
            return jsonify({
                'message': f'Updated {updated_count} suppliers',
                'updated_count': updated_count
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error updating supplier stats: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers/<int:supplier_id>/update-stats', methods=['POST'])
    @login_required
    def update_single_supplier_stats(supplier_id):
        """Update a single supplier's stats for the current shop"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            supplier = Supplier.query.filter_by(id=supplier_id, shop_id=shop_id).first()
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            
            products = Product.query.filter_by(supplier_id=supplier.id, shop_id=shop_id).all()
            supplier.total_products = len(products)
            
            db.session.commit()
            
            return jsonify(supplier.to_dict())
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error updating supplier stats: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers/<int:supplier_id>/products', methods=['GET'])
    @login_required
    def get_supplier_products(supplier_id):
        """Get all products for a specific supplier in the current shop"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            supplier = Supplier.query.filter_by(id=supplier_id, shop_id=shop_id).first()
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            
            products = Product.query.filter_by(supplier_id=supplier_id, shop_id=shop_id).all()
            return jsonify([p.to_dict() for p in products])
            
        except Exception as e:
            logger.error(f"Error fetching supplier products: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers/names', methods=['GET'])
    @login_required
    def get_supplier_names():
        """Get all supplier names for the current shop (for dropdowns)"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            suppliers = Supplier.query.filter_by(shop_id=shop_id, status='Active').order_by(Supplier.name).all()
            return jsonify([{'id': s.id, 'name': s.name, 'item_name': s.item_name} for s in suppliers])
            
        except Exception as e:
            logger.error(f"Error fetching supplier names: {e}")
            return jsonify({'error': str(e)}), 500
    
    # ============ PURCHASE ORDER ROUTES ============
    
    @app.route('/api/purchase-orders', methods=['POST'])
    @login_required
    def create_purchase_order():
        """Create a new purchase order for the current shop"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            data = request.get_json()
            
            if not data or 'supplier_id' not in data:
                return jsonify({'error': 'Supplier ID is required'}), 400
            
            if 'items' not in data or len(data['items']) == 0:
                return jsonify({'error': 'At least one item is required'}), 400
            
            supplier = Supplier.query.filter_by(id=data['supplier_id'], shop_id=shop_id).first()
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            
            subtotal = 0
            items = []
            
            for item_data in data['items']:
                # Check if product_id is provided or if we need to use manual entry
                if 'product_id' in item_data and item_data['product_id']:
                    # Try to find existing product
                    product = Product.query.filter_by(id=item_data['product_id'], shop_id=shop_id).first()
                    if not product:
                        return jsonify({'error': f'Product {item_data["product_id"]} not found'}), 404
                    
                    price = item_data.get('price', product.price)
                    quantity = item_data['quantity']
                    item_total = price * quantity
                    
                    items.append({
                        'product': product,
                        'price': price,
                        'quantity': quantity,
                        'total': item_total
                    })
                else:
                    # Manual entry - use product_name directly
                    product_name = item_data.get('product_name', '')
                    if not product_name:
                        return jsonify({'error': 'Product name is required for manual entry'}), 400
                    
                    # Try to find an existing product by name
                    product = Product.query.filter_by(
                        name=product_name, 
                        shop_id=shop_id
                    ).first()
                    
                    price = item_data.get('price', 0)
                    quantity = item_data.get('quantity', 1)
                    item_total = price * quantity
                    
                    if not product:
                        # Create a new product in the database
                        product = Product(
                            shop_id=shop_id,
                            name=product_name,
                            sku=f"MANUAL-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{len(items)+1}",
                            price=price,
                            stock=0,  # Will be updated when order is received
                            description=f"Added from purchase order",
                            supplier_id=supplier.id
                        )
                        db.session.add(product)
                        db.session.flush()  # Get the product ID
                    
                    items.append({
                        'product': product,
                        'price': price,
                        'quantity': quantity,
                        'total': item_total
                    })
                
                subtotal += items[-1]['total']
            
            tax = data.get('tax', subtotal * 0.16)
            discount = data.get('discount', 0)
            total = subtotal + tax - discount
            
            order = PurchaseOrder(
                shop_id=shop_id,
                supplier_id=data['supplier_id'],
                order_date=datetime.strptime(data.get('order_date', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d').date(),
                subtotal=subtotal,
                tax=tax,
                discount=discount,
                total=total,
                notes=data.get('notes', ''),
                status='Pending',
                created_by=current_user.id if current_user.is_authenticated else None
            )
            
            order.order_number = order.generate_order_number()
            
            db.session.add(order)
            db.session.flush()
            
            # Log initial status
            history = OrderStatusHistory(
                order_id=order.id,
                old_status='',
                new_status='Pending',
                changed_by=current_user.id if current_user.is_authenticated else None,
                notes='Order created'
            )
            db.session.add(history)
            
            for item_data in items:
                product = item_data['product']
                order_item = PurchaseOrderItem(
                    shop_id=shop_id,
                    order_id=order.id,
                    product_id=product.id,
                    product_name=product.name,
                    product_sku=product.sku,
                    quantity=item_data['quantity'],
                    price=item_data['price'],
                    total=item_data['total']
                )
                db.session.add(order_item)
            
            supplier.total_orders += 1
            supplier.last_order_date = datetime.utcnow()
            
            db.session.commit()
            
            logger.info(f"Purchase order created: {order.order_number} for shop {shop_id}")
            
            # Get order dict with shop name
            order_dict = order.to_dict()
            order_dict = enrich_order_with_shop_name(order_dict, order)
            
            return jsonify(order_dict), 201
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error creating purchase order: {e}")
            return jsonify({'error': f'Failed to create purchase order: {str(e)}'}), 500
    
    @app.route('/api/purchase-orders', methods=['GET'])
    @login_required
    def get_purchase_orders():
        """Get all purchase orders for the current shop only"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            query = PurchaseOrder.query.filter_by(shop_id=shop_id)
            
            status = request.args.get('status')
            supplier_id = request.args.get('supplier_id')
            
            if status and status != 'All':
                query = query.filter_by(status=status)
            if supplier_id:
                supplier = Supplier.query.filter_by(id=supplier_id, shop_id=shop_id).first()
                if supplier:
                    query = query.filter_by(supplier_id=supplier_id)
                else:
                    return jsonify({'error': 'Supplier not found'}), 404
            
            orders = query.order_by(PurchaseOrder.created_at.desc()).all()
            
            # Enrich each order with shop name
            result = []
            for order in orders:
                order_dict = order.to_dict()
                order_dict = enrich_order_with_shop_name(order_dict, order)
                result.append(order_dict)
            
            return jsonify(result), 200
            
        except Exception as e:
            logger.error(f"Error fetching purchase orders: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/purchase-orders/<int:order_id>', methods=['GET'])
    @login_required
    def get_purchase_order(order_id):
        """Get a specific purchase order for the current shop"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            order = PurchaseOrder.query.filter_by(id=order_id, shop_id=shop_id).first()
            if not order:
                return jsonify({'error': 'Purchase order not found'}), 404
            
            order_dict = order.to_dict()
            order_dict = enrich_order_with_shop_name(order_dict, order)
            
            return jsonify(order_dict), 200
            
        except Exception as e:
            logger.error(f"Error fetching purchase order: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/purchase-orders/<int:order_id>/status', methods=['PATCH'])
    @login_required
    def update_order_status(order_id):
        """
        Update purchase order status with validation and history logging.
        Supports reverting to previous statuses to fix mistakes.
        """
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            order = PurchaseOrder.query.filter_by(id=order_id, shop_id=shop_id).first()
            if not order:
                return jsonify({'error': 'Purchase order not found'}), 404
            
            data = request.get_json()
            new_status = data.get('status')
            
            if not new_status:
                return jsonify({'error': 'Status is required'}), 400
            
            valid_statuses = ['Pending', 'Ordered', 'Received', 'Cancelled']
            if new_status not in valid_statuses:
                return jsonify({
                    'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
                }), 400
            
            # Validate status transition
            if not validate_status_transition(order.status, new_status):
                available = get_available_status_transitions(order.status)
                if not available:
                    return jsonify({
                        'error': f'Order is already in final state: {order.status}. No further changes allowed.'
                    }), 400
                return jsonify({
                    'error': f'Cannot change status from {order.status} to {new_status}. '
                             f'Allowed transitions: {", ".join(available)}'
                }), 400
            
            old_status = order.status
            
            # ===== INVENTORY MANAGEMENT =====
            # If reverting FROM Received, deduct stock
            if old_status == 'Received' and new_status != 'Received':
                for item in order.items:
                    product = Product.query.filter_by(id=item.product_id, shop_id=shop_id).first()
                    if product:
                        product.stock = max(0, product.stock - item.quantity)
                        product.updated_at = datetime.utcnow()
                        logger.info(f"Stock deducted for product {product.name}: -{item.quantity}")
            
            # If moving TO Received, add stock
            if new_status == 'Received' and old_status != 'Received':
                for item in order.items:
                    product = Product.query.filter_by(id=item.product_id, shop_id=shop_id).first()
                    if product:
                        product.stock += item.quantity
                        product.updated_at = datetime.utcnow()
                        logger.info(f"Stock added for product {product.name}: +{item.quantity}")
            
            # Update order status
            order.status = new_status
            order.updated_at = datetime.utcnow()
            
            # Log status change in history
            history = OrderStatusHistory(
                order_id=order.id,
                old_status=old_status,
                new_status=new_status,
                changed_by=current_user.id if current_user.is_authenticated else None,
                notes=f'Status changed from {old_status} to {new_status}'
            )
            db.session.add(history)
            
            db.session.commit()
            
            logger.info(f"Order {order.order_number} status changed from {old_status} to {new_status} by user {current_user.id}")
            
            # Get order dict with shop name
            order_dict = order.to_dict()
            order_dict = enrich_order_with_shop_name(order_dict, order)
            
            # Get history dict with shop name
            history_dict = history.to_dict()
            history_dict = enrich_history_with_shop_name(history_dict, history)
            
            return jsonify({
                'success': True,
                'message': f'Order status updated from {old_status} to {new_status}',
                'order': order_dict,
                'history': history_dict
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error updating order status: {e}")
            return jsonify({'error': f'Failed to update order status: {str(e)}'}), 500
    
    @app.route('/api/purchase-orders/<int:order_id>/history', methods=['GET'])
    @login_required
    def get_order_history(order_id):
        """Get status history for a specific order"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            order = PurchaseOrder.query.filter_by(id=order_id, shop_id=shop_id).first()
            if not order:
                return jsonify({'error': 'Purchase order not found'}), 404
            
            history = OrderStatusHistory.query.filter_by(order_id=order.id).order_by(
                OrderStatusHistory.created_at.desc()
            ).all()
            
            # Enrich each history entry with shop name
            result = []
            for h in history:
                h_dict = h.to_dict()
                h_dict = enrich_history_with_shop_name(h_dict, h)
                result.append(h_dict)
            
            return jsonify(result), 200
            
        except Exception as e:
            logger.error(f"Error fetching order history: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/purchase-orders/<int:order_id>', methods=['DELETE'])
    @login_required
    def delete_purchase_order(order_id):
        """Delete a purchase order for the current shop (soft delete - mark as cancelled)"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            order = PurchaseOrder.query.filter_by(id=order_id, shop_id=shop_id).first()
            if not order:
                return jsonify({'error': 'Purchase order not found'}), 404
            
            # If order is Received, check if it was a mistake
            if order.status == 'Received':
                # Check if it was recently received (within last hour)
                time_diff = datetime.utcnow() - order.updated_at
                if time_diff.total_seconds() < 3600:  # 1 hour grace period
                    return jsonify({
                        'warning': 'Order was received recently. Consider reverting to Ordered instead.',
                        'confirm': True
                    }), 400
                else:
                    return jsonify({
                        'error': 'Cannot cancel an order that has been received more than 1 hour ago.'
                    }), 400
            
            old_status = order.status
            order.status = 'Cancelled'
            order.updated_at = datetime.utcnow()
            
            # Log cancellation
            history = OrderStatusHistory(
                order_id=order.id,
                old_status=old_status,
                new_status='Cancelled',
                changed_by=current_user.id if current_user.is_authenticated else None,
                notes='Order cancelled'
            )
            db.session.add(history)
            
            db.session.commit()
            
            logger.info(f"Order {order.order_number} cancelled by user {current_user.id}")
            
            return jsonify({'message': 'Purchase order cancelled successfully'}), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error cancelling purchase order: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/purchase-orders/stats', methods=['GET'])
    @login_required
    def get_purchase_order_stats():
        """Get purchase order statistics for the current shop only"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            all_orders = PurchaseOrder.query.filter_by(shop_id=shop_id).all()
            pending = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Pending').count()
            ordered = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Ordered').count()
            received = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Received').count()
            cancelled = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Cancelled').count()
            
            total_amount = sum(o.total for o in all_orders)
            total_orders = len(all_orders)
            
            return jsonify({
                'total': total_orders,
                'pending': pending,
                'ordered': ordered,
                'received': received,
                'cancelled': cancelled,
                'total_amount': float(total_amount)
            })
            
        except Exception as e:
            logger.error(f"Error fetching purchase order stats: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/purchase-orders/status-stats', methods=['GET'])
    @login_required
    def get_order_status_stats():
        """Get statistics about order statuses with recent changes"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            total = PurchaseOrder.query.filter_by(shop_id=shop_id).count()
            pending = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Pending').count()
            ordered = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Ordered').count()
            received = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Received').count()
            cancelled = PurchaseOrder.query.filter_by(shop_id=shop_id, status='Cancelled').count()
            
            # Get recent status changes
            recent_changes = OrderStatusHistory.query.join(
                PurchaseOrder
            ).filter(
                PurchaseOrder.shop_id == shop_id
            ).order_by(
                OrderStatusHistory.created_at.desc()
            ).limit(10).all()
            
            # Enrich history entries with shop name
            enriched_changes = []
            for h in recent_changes:
                h_dict = h.to_dict()
                h_dict = enrich_history_with_shop_name(h_dict, h)
                enriched_changes.append(h_dict)
            
            return jsonify({
                'total': total,
                'pending': pending,
                'ordered': ordered,
                'received': received,
                'cancelled': cancelled,
                'recent_changes': enriched_changes
            }), 200
            
        except Exception as e:
            logger.error(f"Error fetching order status stats: {e}")
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/suppliers/<int:supplier_id>/orders/history', methods=['GET'])
    @login_required
    def get_supplier_order_history(supplier_id):
        """Get all orders with status history for a specific supplier"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            supplier = Supplier.query.filter_by(id=supplier_id, shop_id=shop_id).first()
            if not supplier:
                return jsonify({'error': 'Supplier not found'}), 404
            
            orders = PurchaseOrder.query.filter_by(
                supplier_id=supplier_id,
                shop_id=shop_id
            ).order_by(PurchaseOrder.created_at.desc()).all()
            
            # Enrich each order with shop name
            result = []
            for order in orders:
                order_dict = order.to_dict()
                order_dict = enrich_order_with_shop_name(order_dict, order)
                result.append(order_dict)
            
            return jsonify(result), 200
            
        except Exception as e:
            logger.error(f"Error fetching supplier order history: {e}")
            return jsonify({'error': str(e)}), 500

    # ============ FIX MISSING HISTORIES ============
    
    @app.route('/api/fix-order-histories', methods=['POST'])
    @login_required
    def fix_order_histories():
        """One-time fix to add missing order histories"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            orders = PurchaseOrder.query.filter_by(shop_id=shop_id).all()
            fixed_count = 0
            
            for order in orders:
                history_count = OrderStatusHistory.query.filter_by(order_id=order.id).count()
                if history_count == 0:
                    history = OrderStatusHistory(
                        order_id=order.id,
                        old_status='',
                        new_status=order.status,
                        changed_by=current_user.id if current_user.is_authenticated else None,
                        notes='Order created (history added by fix)',
                        created_at=order.created_at or datetime.utcnow()
                    )
                    db.session.add(history)
                    fixed_count += 1
            
            db.session.commit()
            logger.info(f"Fixed {fixed_count} orders missing history for shop {shop_id}")
            
            return jsonify({
                'success': True,
                'message': f'Fixed {fixed_count} orders missing history',
                'fixed_count': fixed_count
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error fixing order histories: {e}")
            return jsonify({'error': str(e)}), 500

    # ============ ORDER STATUS TRANSITIONS INFO ============
    
    @app.route('/api/order-status-transitions', methods=['GET'])
    @login_required
    def get_status_transitions():
        """Get available status transitions for all statuses"""
        try:
            return jsonify({
                'success': True,
                'transitions': {
                    'Pending': get_available_status_transitions('Pending'),
                    'Ordered': get_available_status_transitions('Ordered'),
                    'Received': get_available_status_transitions('Received'),
                    'Cancelled': get_available_status_transitions('Cancelled')
                }
            }), 200
        except Exception as e:
            logger.error(f"Error fetching status transitions: {e}")
            return jsonify({'error': str(e)}), 500

    # ============ REVERT ORDER STATUS (MISTAKE FIX) ============
    
    @app.route('/api/purchase-orders/<int:order_id>/revert', methods=['POST'])
    @login_required
    def revert_order_status(order_id):
        """Special endpoint to revert a mistaken status change"""
        try:
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            order = PurchaseOrder.query.filter_by(id=order_id, shop_id=shop_id).first()
            if not order:
                return jsonify({'error': 'Purchase order not found'}), 404
            
            data = request.get_json()
            target_status = data.get('target_status')
            
            if not target_status:
                return jsonify({'error': 'Target status is required'}), 400
            
            # Check if reversion is allowed
            if order.status == 'Received':
                if target_status not in ['Ordered', 'Pending']:
                    return jsonify({
                        'error': f'Cannot revert from Received to {target_status}. Allowed: Ordered, Pending'
                    }), 400
            elif order.status == 'Cancelled':
                if target_status not in ['Pending', 'Ordered']:
                    return jsonify({
                        'error': f'Cannot revert from Cancelled to {target_status}. Allowed: Pending, Ordered'
                    }), 400
            elif order.status == 'Ordered':
                if target_status != 'Pending':
                    return jsonify({
                        'error': f'Cannot revert from Ordered to {target_status}. Allowed: Pending'
                    }), 400
            else:
                return jsonify({
                    'error': f'Order is in {order.status} state. Cannot revert.'
                }), 400
            
            # Perform the reversion
            old_status = order.status
            
            # If reverting from Received, deduct stock
            if old_status == 'Received':
                for item in order.items:
                    product = Product.query.filter_by(id=item.product_id, shop_id=shop_id).first()
                    if product:
                        product.stock = max(0, product.stock - item.quantity)
                        product.updated_at = datetime.utcnow()
            
            order.status = target_status
            order.updated_at = datetime.utcnow()
            
            # Log the reversion
            history = OrderStatusHistory(
                order_id=order.id,
                old_status=old_status,
                new_status=target_status,
                changed_by=current_user.id if current_user.is_authenticated else None,
                notes=f'REVERTED from {old_status} to {target_status} (mistake fix)'
            )
            db.session.add(history)
            
            db.session.commit()
            
            logger.info(f"Order {order.order_number} REVERTED from {old_status} to {target_status} by user {current_user.id}")
            
            # Get order dict with shop name
            order_dict = order.to_dict()
            order_dict = enrich_order_with_shop_name(order_dict, order)
            
            # Get history dict with shop name
            history_dict = history.to_dict()
            history_dict = enrich_history_with_shop_name(history_dict, history)
            
            return jsonify({
                'success': True,
                'message': f'Order reverted from {old_status} to {target_status}',
                'order': order_dict,
                'history': history_dict
            }), 200
            
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error reverting order status: {e}")
            return jsonify({'error': str(e)}), 500

    # ============ INVOICE-STYLE PDF EXPORT ============
    
    @app.route('/api/purchase-orders/<int:order_id>/export/pdf', methods=['GET'])
    @login_required
    def export_order_pdf(order_id):
        """Export purchase order as a professional invoice-style PDF"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
            import io
            
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            order = PurchaseOrder.query.filter_by(id=order_id, shop_id=shop_id).first()
            if not order:
                return jsonify({'error': 'Purchase order not found'}), 404
            
            # Get shop details
            shop = Shop.query.filter_by(id=shop_id).first()
            
            # Get created_by_name
            created_by_name = get_shop_name(order.created_by)
            
            # Create PDF buffer
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, 
                                   rightMargin=72, leftMargin=72, 
                                   topMargin=72, bottomMargin=72)
            
            # Styles
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'InvoiceTitle',
                parent=styles['Heading1'],
                fontSize=22,
                textColor=colors.HexColor('#1a56db'),
                alignment=TA_LEFT,
                spaceAfter=2,
                fontName='Helvetica-Bold'
            )
            
            subtitle_style = ParagraphStyle(
                'InvoiceSubtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#6b7280'),
                alignment=TA_LEFT,
                spaceAfter=2,
                fontName='Helvetica'
            )
            
            section_title_style = ParagraphStyle(
                'SectionTitle',
                parent=styles['Heading2'],
                fontSize=11,
                textColor=colors.HexColor('#374151'),
                spaceAfter=4,
                fontName='Helvetica-Bold'
            )
            
            # Build content
            content = []
            
            # ===== HEADER - Left aligned =====
            shop_name = shop.name if shop else 'My Shop'
            content.append(Paragraph(f"<b>{shop_name}</b>", title_style))
            
            # Address
            address_parts = []
            if shop and shop.address:
                address_parts.append(shop.address)
            if address_parts:
                content.append(Paragraph(", ".join(address_parts), subtitle_style))
            
            # Contact
            contact_parts = []
            if shop and hasattr(shop, 'phone') and shop.phone:
                contact_parts.append(f"Tel: {shop.phone}")
            if shop and hasattr(shop, 'email') and shop.email:
                contact_parts.append(f"Email: {shop.email}")
            if contact_parts:
                content.append(Paragraph(" | ".join(contact_parts), subtitle_style))
            
            content.append(Spacer(1, 12))
            
            # ===== ORDER INFO & SUPPLIER INFO - Two column layout =====
            # Left column: Order details
            order_info_data = [
                ['PO Number:', order.order_number],
                ['Date:', order.order_date.strftime('%B %d, %Y') if order.order_date else 'N/A'],
                ['Status:', order.status],
                ['Created By:', created_by_name],
            ]
            
            # Right column: Supplier details
            supplier_info_data = []
            if order.supplier:
                supplier_info_data = [
                    ['Supplier:', order.supplier.name],
                    ['Contact:', order.supplier.contact_person or 'N/A'],
                    ['Email:', order.supplier.email or 'N/A'],
                    ['Phone:', order.supplier.phone or 'N/A'],
                ]
                if order.supplier.address:
                    supplier_info_data.append(['Address:', order.supplier.address])
                if order.supplier.city:
                    supplier_info_data.append(['City:', order.supplier.city])
                if order.supplier.country:
                    supplier_info_data.append(['Country:', order.supplier.country])
            
            # Build rows for two-column table
            max_rows = max(len(order_info_data), len(supplier_info_data)) if supplier_info_data else len(order_info_data)
            combined_data = []
            
            for i in range(max_rows):
                row = []
                # Left column
                if i < len(order_info_data):
                    row.extend([order_info_data[i][0], order_info_data[i][1]])
                else:
                    row.extend(['', ''])
                # Right column
                if supplier_info_data and i < len(supplier_info_data):
                    row.extend([supplier_info_data[i][0], supplier_info_data[i][1]])
                elif supplier_info_data:
                    row.extend(['', ''])
                combined_data.append(row)
            
            # Create the two-column table
            if supplier_info_data:
                info_table = Table(combined_data, colWidths=[80, 120, 80, 120])
                info_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b7280')),
                    ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#6b7280')),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1f2937')),
                    ('TEXTCOLOR', (3, 0), (3, -1), colors.HexColor('#1f2937')),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (2, 0), (2, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('ALIGN', (3, 0), (3, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
            else:
                # Single column if no supplier
                info_table = Table(order_info_data, colWidths=[80, 300])
                info_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b7280')),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1f2937')),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
            
            content.append(info_table)
            content.append(Spacer(1, 15))
            
            # ===== ITEMS TABLE =====
            content.append(Paragraph("Order Items", section_title_style))
            
            # Items table with clean headers
            item_data = [
                ['#', 'Product Description', 'SKU', 'Qty', 'Unit Price', 'Total']
            ]
            
            for idx, item in enumerate(order.items, 1):
                item_data.append([
                    str(idx),
                    item.product_name,
                    item.product_sku,
                    str(item.quantity),
                    f'KES {item.price:,.2f}',
                    f'KES {item.total:,.2f}'
                ])
            
            items_table = Table(item_data, colWidths=[25, 180, 70, 35, 70, 90])
            items_table.setStyle(TableStyle([
                # Header style
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                # Body style
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                # Grid and borders
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
                # Padding
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            content.append(items_table)
            content.append(Spacer(1, 10))
            
            # ===== TOTALS SECTION =====
            totals_data = [
                ['Subtotal:', f'KES {order.subtotal:,.2f}'],
                ['Tax (16%):', f'KES {order.tax:,.2f}'],
                ['Discount:', f'KES {order.discount:,.2f}'],
                ['', ''],
                ['TOTAL:', f'KES {order.total:,.2f}'],
            ]
            
            totals_table = Table(totals_data, colWidths=[120, 120])
            totals_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b7280')),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1f2937')),
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('TEXTCOLOR', (1, -1), (1, -1), colors.HexColor('#1a56db')),
            ]))
            
            totals_wrapper = Table([[totals_table]], colWidths=[400])
            totals_wrapper.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ]))
            content.append(totals_wrapper)
            
            # ===== NOTES =====
            if order.notes:
                content.append(Spacer(1, 10))
                content.append(Paragraph("Notes:", section_title_style))
                content.append(Paragraph(order.notes, styles['Normal']))
            
            # ===== FOOTER =====
            content.append(Spacer(1, 30))
            
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#9ca3af'),
                alignment=TA_CENTER,
            )
            content.append(Paragraph("Thank you for your business!", footer_style))
            
            terms_style = ParagraphStyle(
                'Terms',
                parent=styles['Normal'],
                fontSize=7,
                textColor=colors.HexColor('#d1d5db'),
                alignment=TA_CENTER,
            )
            content.append(Paragraph("Returns are accepted within 30 days of the purchase date.", terms_style))
            
            # Build PDF
            doc.build(content)
            
            pdf_data = buffer.getvalue()
            buffer.close()
            
            response = current_app.response_class(
                pdf_data,
                mimetype='application/pdf'
            )
            response.headers['Content-Disposition'] = f'attachment; filename=INVOICE-{order.order_number}.pdf'
            return response
            
        except Exception as e:
            logger.error(f"Error exporting order as PDF: {e}")
            return jsonify({'error': str(e)}), 500

    # ============ EXCEL EXPORT ============
    
    @app.route('/api/purchase-orders/<int:order_id>/export/excel', methods=['GET'])
    @login_required
    def export_order_excel(order_id):
        """Export purchase order as Excel with shop name instead of User ID"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io
            
            shop_id = get_current_shop_id()
            if not shop_id:
                return jsonify({'error': 'Shop not found'}), 401
            
            order = PurchaseOrder.query.filter_by(id=order_id, shop_id=shop_id).first()
            if not order:
                return jsonify({'error': 'Purchase order not found'}), 404
            
            # Get created_by_name
            created_by_name = get_shop_name(order.created_by)
            
            # Get shop details
            shop = Shop.query.filter_by(id=shop_id).first()
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"PO-{order.order_number}"
            
            # Styles
            title_font = Font(bold=True, size=14, color="1a56db")
            header_font = Font(bold=True, size=10, color="374151")
            header_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            total_font = Font(bold=True, size=10)
            grand_total_font = Font(bold=True, size=12, color="1a56db")
            border = Border(
                left=Side(style='thin', color="d1d5db"),
                right=Side(style='thin', color="d1d5db"),
                top=Side(style='thin', color="d1d5db"),
                bottom=Side(style='thin', color="d1d5db")
            )
            
            # ===== HEADER =====
            ws.merge_cells('A1:F1')
            ws['A1'] = shop.name if shop else 'My Shop'
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
            
            if shop:
                row = 2
                address_parts = []
                if shop.address:
                    address_parts.append(shop.address)
                
                if address_parts:
                    ws.merge_cells(f'A{row}:F{row}')
                    ws[f'A{row}'] = ", ".join(address_parts)
                    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
                    ws[f'A{row}'].font = Font(size=9, color="6b7280")
                    row += 1
                
                contact_parts = []
                if hasattr(shop, 'phone') and shop.phone:
                    contact_parts.append(f"Tel: {shop.phone}")
                if hasattr(shop, 'email') and shop.email:
                    contact_parts.append(f"Email: {shop.email}")
                if contact_parts:
                    ws.merge_cells(f'A{row}:F{row}')
                    ws[f'A{row}'] = " | ".join(contact_parts)
                    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
                    ws[f'A{row}'].font = Font(size=9, color="6b7280")
            
            # ===== ORDER INFO & SUPPLIER INFO - Two column =====
            row = 5
            
            # Left column: Order Info
            order_labels = ['PO Number:', 'Date:', 'Status:', 'Created By:']
            order_values = [
                order.order_number,
                order.order_date.strftime('%B %d, %Y') if order.order_date else 'N/A',
                order.status,
                created_by_name
            ]
            
            # Right column: Supplier Info
            supplier_labels = []
            supplier_values = []
            if order.supplier:
                supplier_labels = ['Supplier:', 'Contact:', 'Email:', 'Phone:']
                supplier_values = [
                    order.supplier.name,
                    order.supplier.contact_person or 'N/A',
                    order.supplier.email or 'N/A',
                    order.supplier.phone or 'N/A'
                ]
                if order.supplier.address:
                    supplier_labels.append('Address:')
                    supplier_values.append(order.supplier.address)
                if order.supplier.city:
                    supplier_labels.append('City:')
                    supplier_values.append(order.supplier.city)
                if order.supplier.country:
                    supplier_labels.append('Country:')
                    supplier_values.append(order.supplier.country)
            
            # Write two-column data
            max_rows = max(len(order_labels), len(supplier_labels))
            for i in range(max_rows):
                # Left column
                if i < len(order_labels):
                    ws.cell(row=row, column=1, value=order_labels[i]).font = Font(bold=True, color="6b7280", size=8)
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
                    ws.cell(row=row, column=2, value=order_values[i]).font = Font(size=8)
                
                # Right column
                if i < len(supplier_labels):
                    ws.cell(row=row, column=3, value=supplier_labels[i]).font = Font(bold=True, color="6b7280", size=8)
                    ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
                    ws.cell(row=row, column=4, value=supplier_values[i]).font = Font(size=8)
                
                row += 1
            
            # ===== ITEMS TABLE =====
            row += 2
            headers = ['#', 'Product', 'SKU', 'Quantity', 'Price', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            row += 1
            for idx, item in enumerate(order.items, 1):
                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=2, value=item.product_name).border = border
                ws.cell(row=row, column=3, value=item.product_sku).border = border
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=4, value=item.quantity).border = border
                ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=row, column=5, value=f"KES {item.price:,.2f}").border = border
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=6, value=f"KES {item.total:,.2f}").border = border
                ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
                row += 1
            
            # ===== TOTALS =====
            row += 1
            totals = [
                ('Subtotal:', f"KES {order.subtotal:,.2f}"),
                ('Tax (16%):', f"KES {order.tax:,.2f}"),
                ('Discount:', f"KES {order.discount:,.2f}"),
                ('', ''),
                ('TOTAL:', f"KES {order.total:,.2f}"),
            ]
            
            for label, value in totals:
                ws.cell(row=row, column=5, value=label).font = Font(bold=True, size=9)
                ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
                ws.cell(row=row, column=6, value=value)
                if label == 'TOTAL:':
                    ws.cell(row=row, column=6).font = grand_total_font
                else:
                    ws.cell(row=row, column=6).font = total_font
                row += 1
            
            # ===== NOTES =====
            if order.notes:
                row += 1
                ws.cell(row=row, column=1, value="Notes:").font = Font(bold=True, size=9)
                ws.merge_cells(f'B{row}:F{row}')
                ws.cell(row=row, column=2, value=order.notes).font = Font(size=9)
            
            # ===== FOOTER =====
            row += 2
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "Thank you for your business!"
            ws[f'A{row}'].alignment = Alignment(horizontal="center")
            ws[f'A{row}'].font = Font(size=8, color="9ca3af")
            
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "Returns are accepted within 30 days of the purchase date."
            ws[f'A{row}'].alignment = Alignment(horizontal="center")
            ws[f'A{row}'].font = Font(size=7, color="d1d5db")
            
            # Auto-adjust column widths
            for col in range(1, 7):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].auto_size = True
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_data = buffer.getvalue()
            buffer.close()
            
            response = current_app.response_class(
                excel_data,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response.headers['Content-Disposition'] = f'attachment; filename=INVOICE-{order.order_number}.xlsx'
            return response
            
        except Exception as e:
            logger.error(f"Error exporting order as Excel: {e}")
            return jsonify({'error': str(e)}), 500